#!/usr/bin/env python3
"""
Aggregate benchmark JSON results into normalized Excel workbook.

Schema:
- Benchmarks: One row per benchmark run (machine/GPU metadata)
- Summary: Test statistics linked by benchmark_id
- HW_RT_Frames, Compute_Frames, Fragment_Frames: Per-frame data by pipeline
- Cross_Machine: Comparison across different machines/GPUs

Directory Structure for Multi-Tester Data:
    data/benchmarks/
        benchmark_001/          <- Individual benchmark folders
            *.json
            debug_images/
        benchmark_002/
        benchmark_001.zip       <- Compressed archives for transport

Usage:
    python aggregate_results.py [--output PATH] [--cleanup]
    python aggregate_results.py --pack benchmark_001    # Create ZIP from folder
    python aggregate_results.py --unpack benchmark.zip  # Extract ZIP to folder
    python aggregate_results.py --process-all           # Process all benchmarks in data/benchmarks/
"""

import json
import argparse
import uuid
import os
import shutil
import socket
import zipfile
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from chart_config import PROJECT_ROOT, DATA_DIR, BENCHMARK_RESULTS_DIR, EXCEL_FILE

# Multi-tester benchmark directory
BENCHMARKS_DIR = DATA_DIR / "benchmarks"

# Pipeline name mapping for sheet names
PIPELINE_SHEET_NAMES = {
    'hardware_rt': 'HW_RT_Frames',
    'compute': 'Compute_Frames',
    'fragment': 'Fragment_Frames',
}


def get_machine_name() -> str:
    """Get machine name from environment, config, or hostname."""
    # Check environment variable first
    if 'VIXEN_MACHINE_NAME' in os.environ:
        return os.environ['VIXEN_MACHINE_NAME']

    # Check benchmark_config.json for machine_name
    config_path = PROJECT_ROOT / 'application' / 'benchmark' / 'benchmark_config.json'
    if config_path.exists():
        try:
            with open(config_path) as f:
                config = json.load(f)
                if 'suite' in config and config['suite'].get('machine_name'):
                    return config['suite']['machine_name']
        except (json.JSONDecodeError, IOError):
            pass

    # Fall back to hostname
    return socket.gethostname()


def generate_benchmark_id() -> str:
    """Generate a unique benchmark ID."""
    return str(uuid.uuid4())[:8]


def load_benchmark_results(results_dir: Path) -> list[dict]:
    """Load all JSON benchmark results from directory."""
    results = []

    json_files = sorted(results_dir.glob("*.json"))
    print(f"Found {len(json_files)} JSON files in {results_dir}")

    for json_file in json_files:
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
                data['_source_file'] = json_file.name
                results.append(data)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: Could not load {json_file}: {e}")

    return results


def extract_benchmark_metadata(results: list[dict], benchmark_id: str, machine_name: str) -> pd.DataFrame:
    """Extract benchmark run metadata (one row per benchmark run)."""
    if not results:
        return pd.DataFrame()

    # Get device info from first result (same for all in a run)
    device = results[0].get('device', {})

    # Find the earliest timestamp
    timestamps = [r.get('timestamp', '') for r in results if r.get('timestamp')]
    run_date = min(timestamps) if timestamps else datetime.now().isoformat()

    row = {
        'benchmark_id': benchmark_id,
        'machine_name': machine_name,
        'gpu_name': device.get('gpu', 'unknown'),
        'gpu_driver': device.get('driver', 'unknown'),
        'vram_gb': device.get('vram_gb', 0),
        'run_date': run_date,
        'total_tests': len(results),
        'notes': '',
    }

    return pd.DataFrame([row])


def extract_summary_data(results: list[dict], benchmark_id: str) -> pd.DataFrame:
    """Extract summary statistics with benchmark_id reference."""
    rows = []

    for result in results:
        config = result.get('configuration', {})
        stats = result.get('statistics', {})
        pipeline = config.get('pipeline', 'unknown')

        # Skip unknown pipelines/scenes
        if pipeline == 'unknown' or config.get('scene_type', 'unknown') == 'unknown':
            continue

        row = {
            'benchmark_id': benchmark_id,
            'test_id': result.get('test_id', 'unknown'),
            'pipeline': pipeline,
            'resolution': config.get('resolution', 0),
            'scene': config.get('scene_type', 'unknown'),
            'shader': config.get('shader', 'unknown'),
            'screen_width': config.get('screen_width', 0),
            'screen_height': config.get('screen_height', 0),
            'fps_mean': stats.get('fps_mean', 0),
            'frame_time_mean_ms': stats.get('frame_time_mean', 0),
            'frame_time_p99_ms': stats.get('frame_time_p99', 0),
            'frame_time_stddev': stats.get('frame_time_stddev', 0),
            'bandwidth_mean_gbps': stats.get('bandwidth_mean', 0),
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['pipeline', 'resolution', 'scene'])
    return df


def extract_frame_data_by_pipeline(results: list[dict], benchmark_id: str) -> dict[str, pd.DataFrame]:
    """Extract per-frame data separated by pipeline type.

    Note: Frame at midpoint (warmup + measurement/2) is filtered out because
    FrameCapture takes a debug screenshot there, causing artificial spikes.
    """
    pipeline_frames = {name: [] for name in PIPELINE_SHEET_NAMES.values()}

    for result in results:
        config = result.get('configuration', {})
        pipeline = config.get('pipeline', 'unknown')
        test_id = result.get('test_id', 'unknown')
        frames = result.get('frames', [])

        # Skip unknown pipeline
        if pipeline == 'unknown' or pipeline not in PIPELINE_SHEET_NAMES:
            continue

        sheet_name = PIPELINE_SHEET_NAMES[pipeline]

        # Calculate capture frame to filter (midpoint of measurement frames)
        # Default: 50 warmup + 50 (half of 100 measurement) = frame 100 in raw numbering
        # But frame_num in JSON is 0-indexed from measurement start, so midpoint = 50
        measurement_frames = len(frames)
        capture_frame = measurement_frames // 2  # Midpoint where debug capture occurs

        for frame in frames:
            frame_num = frame.get('frame_num', 0)

            # Skip the capture frame (causes artificial spike due to GPU readback)
            if frame_num == capture_frame:
                continue

            row = {
                'benchmark_id': benchmark_id,
                'test_id': test_id,
                'frame': frame_num,
                'frame_time_ms': frame.get('frame_time_ms', 0),
                'fps': frame.get('fps', 0),
                'bandwidth_read_gbps': frame.get('bandwidth_read_gbps', 0),
                'bandwidth_write_gbps': frame.get('bandwidth_write_gbps', 0),
                'vram_mb': frame.get('vram_mb', 0),
                'avg_voxels_per_ray': frame.get('avg_voxels_per_ray', 0),
                'ray_throughput_mrays': frame.get('ray_throughput_mrays', 0),
            }
            pipeline_frames[sheet_name].append(row)

    return {name: pd.DataFrame(rows) for name, rows in pipeline_frames.items()}


def extract_cross_machine_data(wb: Workbook) -> pd.DataFrame:
    """Generate cross-machine comparison from existing Summary data."""
    if 'Summary' not in wb.sheetnames:
        return pd.DataFrame()

    # Read existing summary data
    ws = wb['Summary']
    data = list(ws.values)
    if len(data) < 2:
        return pd.DataFrame()

    headers = data[0]
    rows = data[1:]
    summary_df = pd.DataFrame(rows, columns=headers)

    if summary_df.empty or 'benchmark_id' not in summary_df.columns:
        return pd.DataFrame()

    # Read benchmarks sheet for machine/GPU info
    if 'Benchmarks' not in wb.sheetnames:
        return pd.DataFrame()

    ws_bench = wb['Benchmarks']
    bench_data = list(ws_bench.values)
    if len(bench_data) < 2:
        return pd.DataFrame()

    bench_headers = bench_data[0]
    bench_rows = bench_data[1:]
    bench_df = pd.DataFrame(bench_rows, columns=bench_headers)

    # Merge to get GPU info
    merged = summary_df.merge(bench_df[['benchmark_id', 'machine_name', 'gpu_name']], on='benchmark_id', how='left')

    # Group by GPU and pipeline for comparison
    if 'fps_mean' in merged.columns and 'pipeline' in merged.columns:
        comparison = merged.groupby(['gpu_name', 'pipeline']).agg({
            'fps_mean': ['mean', 'min', 'max'],
            'frame_time_mean_ms': 'mean',
            'test_id': 'count',
        }).round(2)

        comparison.columns = ['avg_fps', 'min_fps', 'max_fps', 'avg_frame_time_ms', 'test_count']
        return comparison.reset_index()

    return pd.DataFrame()


def style_header(ws, row: int, num_cols: int):
    """Apply header styling to a worksheet row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def auto_width_columns(ws):
    """Auto-width all columns in a worksheet."""
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)


def append_to_sheet(ws, df: pd.DataFrame, include_header: bool = False):
    """Append dataframe rows to existing worksheet."""
    if df.empty:
        return

    start_row = ws.max_row + 1 if ws.max_row > 1 or not include_header else 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=include_header), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def create_or_update_workbook(results: list[dict], output_path: Path, benchmark_id: str, machine_name: str):
    """Create new or append to existing Excel workbook."""

    # Check if workbook exists
    if output_path.exists():
        print(f"Appending to existing workbook: {output_path}")
        wb = load_workbook(output_path)
        is_new = False
    else:
        print(f"Creating new workbook: {output_path}")
        wb = Workbook()
        wb.remove(wb.active)
        is_new = True

    # Extract data
    benchmark_df = extract_benchmark_metadata(results, benchmark_id, machine_name)
    summary_df = extract_summary_data(results, benchmark_id)
    frame_dfs = extract_frame_data_by_pipeline(results, benchmark_id)

    # Sheet 1: Benchmarks (metadata)
    if 'Benchmarks' not in wb.sheetnames:
        ws_bench = wb.create_sheet("Benchmarks", 0)
        for r_idx, row in enumerate(dataframe_to_rows(benchmark_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_bench.cell(row=r_idx, column=c_idx, value=value)
        style_header(ws_bench, 1, len(benchmark_df.columns))
        auto_width_columns(ws_bench)
    else:
        ws_bench = wb['Benchmarks']
        append_to_sheet(ws_bench, benchmark_df, include_header=False)

    # Sheet 2: Summary
    if 'Summary' not in wb.sheetnames:
        ws_summary = wb.create_sheet("Summary")
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
        if not summary_df.empty:
            style_header(ws_summary, 1, len(summary_df.columns))
            auto_width_columns(ws_summary)
    else:
        ws_summary = wb['Summary']
        append_to_sheet(ws_summary, summary_df, include_header=False)

    # Sheets 3-5: Frame data by pipeline
    for sheet_name, frame_df in frame_dfs.items():
        if frame_df.empty:
            continue

        if sheet_name not in wb.sheetnames:
            ws_frames = wb.create_sheet(sheet_name)
            for r_idx, row in enumerate(dataframe_to_rows(frame_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_frames.cell(row=r_idx, column=c_idx, value=value)
            style_header(ws_frames, 1, len(frame_df.columns))
            auto_width_columns(ws_frames)
        else:
            ws_frames = wb[sheet_name]
            append_to_sheet(ws_frames, frame_df, include_header=False)

    # Sheet 6: Cross_Machine (regenerate each time)
    if 'Cross_Machine' in wb.sheetnames:
        del wb['Cross_Machine']

    cross_df = extract_cross_machine_data(wb)
    if not cross_df.empty:
        ws_cross = wb.create_sheet("Cross_Machine")
        for r_idx, row in enumerate(dataframe_to_rows(cross_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_cross.cell(row=r_idx, column=c_idx, value=value)
        style_header(ws_cross, 1, len(cross_df.columns))
        auto_width_columns(ws_cross)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel workbook saved: {output_path}")


def cleanup_transient_files(results_dir: Path):
    """Delete JSON files and debug_images after successful aggregation."""
    # Delete JSON files
    json_files = list(results_dir.glob("*.json"))
    for json_file in json_files:
        try:
            json_file.unlink()
            print(f"Deleted: {json_file.name}")
        except IOError as e:
            print(f"Warning: Could not delete {json_file}: {e}")

    # Delete debug_images directory
    debug_images_dir = results_dir / "debug_images"
    if debug_images_dir.exists():
        try:
            shutil.rmtree(debug_images_dir)
            print(f"Deleted: debug_images/")
        except IOError as e:
            print(f"Warning: Could not delete debug_images: {e}")


def pack_benchmark(source_dir: Path, output_zip: Path | None = None) -> Path:
    """
    Pack a benchmark folder into a ZIP archive for transport.

    Args:
        source_dir: Path to benchmark folder containing JSON files
        output_zip: Optional output path (default: same name with .zip extension)

    Returns:
        Path to created ZIP file
    """
    if not source_dir.exists():
        raise FileNotFoundError(f"Source directory not found: {source_dir}")

    if output_zip is None:
        output_zip = source_dir.with_suffix('.zip')

    print(f"Packing {source_dir} -> {output_zip}")

    with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file_path in source_dir.rglob('*'):
            if file_path.is_file():
                arcname = file_path.relative_to(source_dir.parent)
                zf.write(file_path, arcname)
                print(f"  Added: {arcname}")

    # Calculate compression ratio
    original_size = sum(f.stat().st_size for f in source_dir.rglob('*') if f.is_file())
    compressed_size = output_zip.stat().st_size
    ratio = original_size / compressed_size if compressed_size > 0 else 0

    print(f"Packed: {original_size / 1024:.1f} KB -> {compressed_size / 1024:.1f} KB ({ratio:.1f}x compression)")
    return output_zip


def unpack_benchmark(zip_path: Path, output_dir: Path | None = None) -> Path:
    """
    Unpack a benchmark ZIP archive.

    Args:
        zip_path: Path to ZIP file
        output_dir: Optional output directory (default: same location as ZIP)

    Returns:
        Path to extracted directory
    """
    if not zip_path.exists():
        raise FileNotFoundError(f"ZIP file not found: {zip_path}")

    if output_dir is None:
        output_dir = zip_path.parent

    print(f"Unpacking {zip_path} -> {output_dir}")

    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(output_dir)
        # Get the root folder name from the archive
        root_folders = set(Path(n).parts[0] for n in zf.namelist() if '/' in n or '\\' in n)
        if root_folders:
            extracted_dir = output_dir / list(root_folders)[0]
        else:
            extracted_dir = output_dir

    print(f"Extracted to: {extracted_dir}")
    return extracted_dir


def get_benchmark_folders() -> list[Path]:
    """Get all benchmark folders in data/benchmarks/."""
    BENCHMARKS_DIR.mkdir(parents=True, exist_ok=True)

    folders = []
    for item in BENCHMARKS_DIR.iterdir():
        if item.is_dir() and list(item.glob("*.json")):
            folders.append(item)

    return sorted(folders)


def process_all_benchmarks(output_path: Path, cleanup: bool = False) -> int:
    """
    Process all benchmark folders in data/benchmarks/.

    Args:
        output_path: Path to output Excel file
        cleanup: Whether to delete JSON files after aggregation

    Returns:
        Number of benchmarks processed
    """
    folders = get_benchmark_folders()

    if not folders:
        print(f"No benchmark folders found in {BENCHMARKS_DIR}")
        print("Expected structure: data/benchmarks/benchmark_001/*.json")
        return 0

    print(f"Found {len(folders)} benchmark folder(s):")
    for folder in folders:
        json_count = len(list(folder.glob("*.json")))
        print(f"  - {folder.name}: {json_count} JSON files")

    processed = 0
    for folder in folders:
        print(f"\n{'='*60}")
        print(f"Processing: {folder.name}")
        print('='*60)

        results = load_benchmark_results(folder)
        if not results:
            print(f"No valid results in {folder.name}, skipping")
            continue

        # Use folder name as part of benchmark ID for traceability
        benchmark_id = f"{folder.name[:8]}_{generate_benchmark_id()}"
        machine_name = get_machine_name()

        print(f"Benchmark ID: {benchmark_id}")
        print(f"Machine: {machine_name}")

        create_or_update_workbook(results, output_path, benchmark_id, machine_name)
        print(f"Aggregated {len(results)} results from {folder.name}")

        if cleanup:
            print(f"\nCleaning up {folder.name}...")
            cleanup_transient_files(folder)

        processed += 1

    return processed


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate benchmark results to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process single benchmark_results/ directory
  python aggregate_results.py

  # Process all benchmark folders in data/benchmarks/
  python aggregate_results.py --process-all

  # Pack a benchmark folder for sharing
  python aggregate_results.py --pack benchmark_results/

  # Unpack received benchmark archive
  python aggregate_results.py --unpack benchmark_001.zip

  # Process with cleanup
  python aggregate_results.py --process-all --cleanup
        """
    )
    parser.add_argument('--output', '-o', type=Path, default=EXCEL_FILE,
                        help=f"Output Excel file (default: {EXCEL_FILE})")
    parser.add_argument('--input', '-i', type=Path, default=BENCHMARK_RESULTS_DIR,
                        help=f"Input directory with JSON results (default: {BENCHMARK_RESULTS_DIR})")
    parser.add_argument('--cleanup', action='store_true',
                        help="Delete JSON files and debug_images after aggregation")
    parser.add_argument('--machine-name', type=str, default=None,
                        help="Override machine name (default: from env/config/hostname)")

    # New multi-tester options
    parser.add_argument('--process-all', action='store_true',
                        help=f"Process all benchmark folders in {BENCHMARKS_DIR}")
    parser.add_argument('--pack', type=Path, metavar='DIR',
                        help="Pack a benchmark folder into ZIP for transport")
    parser.add_argument('--unpack', type=Path, metavar='ZIP',
                        help="Unpack a benchmark ZIP archive")
    parser.add_argument('--list', '-l', action='store_true',
                        help="List available benchmark folders")

    args = parser.parse_args()

    # Handle --pack command
    if args.pack:
        try:
            zip_path = pack_benchmark(args.pack)
            print(f"\nReady for transport: {zip_path}")
            return 0
        except FileNotFoundError as e:
            print(f"Error: {e}")
            return 1

    # Handle --unpack command
    if args.unpack:
        try:
            extracted = unpack_benchmark(args.unpack, BENCHMARKS_DIR)
            print(f"\nReady for processing: {extracted}")
            print(f"Run: python aggregate_results.py --process-all")
            return 0
        except FileNotFoundError as e:
            print(f"Error: {e}")
            return 1

    # Handle --list command
    if args.list:
        folders = get_benchmark_folders()
        if not folders:
            print(f"No benchmark folders found in {BENCHMARKS_DIR}")
        else:
            print(f"Benchmark folders in {BENCHMARKS_DIR}:")
            for folder in folders:
                json_count = len(list(folder.glob("*.json")))
                print(f"  - {folder.name}: {json_count} JSON files")
        return 0

    # Handle --process-all command
    if args.process_all:
        processed = process_all_benchmarks(args.output, args.cleanup)
        print(f"\nProcessed {processed} benchmark folder(s)")
        return 0 if processed > 0 else 1

    # Default: process single input directory
    if not args.input.exists():
        print(f"Error: Input directory not found: {args.input}")
        return 1

    results = load_benchmark_results(args.input)

    if not results:
        print("No benchmark results found to aggregate.")
        return 1

    # Generate benchmark ID and get machine name
    benchmark_id = generate_benchmark_id()
    machine_name = args.machine_name or get_machine_name()

    print(f"Benchmark ID: {benchmark_id}")
    print(f"Machine: {machine_name}")

    create_or_update_workbook(results, args.output, benchmark_id, machine_name)
    print(f"Aggregated {len(results)} benchmark results")

    if args.cleanup:
        print("\nCleaning up transient files...")
        cleanup_transient_files(args.input)

    return 0


if __name__ == "__main__":
    exit(main())
